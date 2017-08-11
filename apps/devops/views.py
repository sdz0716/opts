from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse, HttpResponseRedirect, Http404, JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from apps.devops.models import serverList
from django.db.models import Count, Sum
from django.core.paginator import Paginator
import json
import datetime
from apps.devops import forms
from django.contrib import messages
from apps.devops import models
import openpyxl

@login_required()
def tables(request):
    return render(request, 'devops/tables.html')

@login_required()
def serverInfo(request):
    if request.method == 'GET':
        # print(request.GET)
        # print(request.POST.getlist('btSelectItem')) #获取checkbox信息
        limit = request.GET.get('limit')
        offset = request.GET.get('offset')  # how many items in total in the DB
        search = request.GET.get('search')
        sort_column = request.GET.get('sort')
        order = request.GET.get('order')
        if search:
            all_records = serverList.objects.filter(instance=search, hostname=search, ip=search, hostcomputer=search, cpucore=search, memory=search, system=search, systemdisk=search, datadisk=search, user=search, use=search, remark=search, id=search)
        else:
            all_records = serverList.objects.all()

        if sort_column:
            if sort_column in ['instance', 'hostname', 'ip', 'hostcomputer', 'cpucore', 'memory', 'system', 'systemdisk', 'datadisk', 'user', 'use', 'remark', 'alter_time', 'id']:
                if order == 'desc':
                    sort_column = '-%s' % sort_column
                all_records = serverList.objects.all().order_by(sort_column)

        all_records_count = all_records.count()
        if not offset:
            offset = 0
        if not limit:
            limit = 10
        pageinator = Paginator(all_records, limit)
        page = int(int(offset) / int(limit) + 1)
        response_data = {'total': all_records_count, 'rows': []}
        for ser in pageinator.page(page):
            response_data['rows'].append({
                'id': ser.id if ser.id else '',
                'instance': ser.instance if ser.instance else '',
                'hostname': ser.hostname if ser.hostname else '',
                'ip': ser.ip if ser.ip else '',
                'hostcomputer': ser.hostcomputer if ser.hostcomputer else '',
                'cpucore': ser.cpucore if ser.cpucore else '',
                'memory': ser.memory if ser.memory else '',
                'system': ser.system if ser.system else '',
                'systemdisk': ser.systemdisk if ser.systemdisk else '',
                'datadisk': ser.datadisk if ser.datadisk else '',
                'user': ser.user if ser.user else '',
                'use': ser.use if ser.use else '',
                'remark': ser.remark if ser.remark else '',
                'alter_time': ser.alter_time.strftime('%Y-%m-%d %H:%M') if ser.alter_time else '',
                'create_time': ser.create_time.strftime('%Y-%m-%d %H:%M') if ser.create_time else '',
            })
    # return HttpResponse(json.dumps(response_data))
    #     print(json.dumps(response_data))
        return JsonResponse(response_data)



@login_required()
def addServerInfo(request):
    # checklist = request.POST.getlist('btSelectItem')
    # print(checklist)
    if request.method == 'POST':
        # print(request.POST)
        form = forms.serverInfoForm(request.POST)
        if form.is_valid():
            form.save()
        return HttpResponseRedirect('/devops/')
    else:
        form = forms.serverInfoForm(initial={'alter_time': datetime.datetime.now()})
    return render(request, 'devops/addserverinfo.html', {'serverInfo': form})

@login_required()
def modifyServerInfo(request, selectd_id):
    obj = models.serverList.objects.get(pk=selectd_id)
    if request.method == 'POST':
        form = forms.serverInfoForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            # baseUrl = '/'.join(request.path.split('/')[:-2])
            # print(baseUrl)
            return HttpResponseRedirect('/devops/')
    if request.method == 'GET':
        form = forms.serverInfoForm(instance=obj)
        # print(form)
    return render(request, 'devops/modifyserverinfo.html', {'serverInfo': form})

@login_required()
def deleteServerInfo(request):
    print(request.POST.getlist('delList[]'))
    for i in request.POST.getlist('delList[]'):
        i = int(i)
        models.serverList.objects.filter(id=i).delete()
    return HttpResponse('success')

@login_required()
def importServerInfo(request):
    if request.method == 'POST':
        myfile = request.FILES.get('file')
        wb = openpyxl.load_workbook(myfile)
        sheet = wb.get_active_sheet()
        serverDict = {}
        try:
            for r in range(2, sheet.max_row+1):
                serverDict['instance'] = sheet.cell(row=r, column=1).value
                serverDict['hostname'] = sheet.cell(row=r, column=2).value
                serverDict['ip'] = sheet.cell(row=r, column=3).value
                serverDict['hostcomputer'] = sheet.cell(row=r, column=4).value
                serverDict['cpucore'] = sheet.cell(row=r, column=5).value
                serverDict['memory'] = sheet.cell(row=r, column=6).value
                serverDict['system'] = sheet.cell(row=r, column=7).value
                serverDict['systemdisk'] = sheet.cell(row=r, column=8).value
                serverDict['datadisk'] = sheet.cell(row=r, column=9).value
                serverDict['user'] = sheet.cell(row=r, column=10).value
                serverDict['use'] = sheet.cell(row=r, column=11).value
                serverDict['remark'] = sheet.cell(row=r, column=12).value
                models.serverList.objects.create(**serverDict)
        except:
            return HttpResponse('<script>alert("导入失败，请检查是否有重复数据");window.history.back(-1);</script>')
        return HttpResponseRedirect('tables.html')
    else:
        return render(request, 'devops/importserverinfo.html')